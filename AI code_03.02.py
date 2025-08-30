from abbrevia import abbreviation
import PyPDF2
import re
from datetime import date  # for date comparison
from openpyxl.styles import PatternFill

# Define a function to extract text from a PDF file
def extract_text_from_pdf(file):
    pdf_reader = PyPDF2.PdfReader(file)
    text = ''
    for page_num in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[page_num]
        text += page.extract_text()
    return text

# Read the PDF file and extract its text
with open('periods/1.03.2024.pdf', 'rb') as f:
    pdf_text = extract_text_from_pdf(f)

Match_amount_description = re.findall(r'([-+]?\d{1,3}(?:[ ,]\d{3})*(?:\.\d{2})?) ₽ [^₽]*₽ (.*[^..]*)', pdf_text)
matches_date = re.findall(r'(\d{2}\.\d{2}\.\d{4})\s\d{2}:\d{2}[+-]', pdf_text)

# Initialize the dictionary structure
dic = {}

from_date = "01.02.2024".split('.')  # enter the start date here
from_date = date(int(from_date[2]), int(from_date[1]), int(from_date[0]))

upto_date = "01.03.2024".split('.')  # enter the end date here
upto_date = date(int(upto_date[2]), int(upto_date[1]), int(upto_date[0]))

for match in range(len(matches_date)):
    Match_amount_description_string = Match_amount_description[match]
    amount_string = float(Match_amount_description_string[0].replace(',', '').replace(' ', ''))
    description_string = abbreviation(Match_amount_description_string[1][:-3])
    date_string = matches_date[match]
    date_int = date_string.split('.')
    date_int = date(int(date_int[2]), int(date_int[1]), int(date_int[0]))

    # Check if the date is within the specified range
    if from_date <= date_int < upto_date:
        if description_string not in dic:
            dic[description_string] = {'amountsPlus': [], 'amountsMinus': [], 'totalminus': 0, 'dates': []}

        if amount_string < 0:
            dic[description_string]['amountsMinus'].append(amount_string)
            dic[description_string]['totalminus'] += amount_string
        else:
            dic[description_string]['amountsPlus'].append(amount_string)

        dic[description_string]['dates'].append(date_string)

print(dic)




# ---------------------Запись в файл--------------------------#

from openpyxl.writer.excel import save_workbook
from openpyxl import *

page = 0

FILE_NAME = 'excel.xlsx'

try:
    wb = load_workbook(FILE_NAME)
except:
    wb = Workbook()

ws = wb.create_sheet(str(from_date) + ' - ' + str(upto_date))

num = 2
_cell = ws['A1']
_cell2 = ws['B1']
_cell3 = ws['C1']
_cell4 = ws['D1']
_cell5 = ws['E1']

_cell.value = 'Описание'
_cell2.value = 'Сумма Plus'
_cell3.value = 'Сумма Minus'
_cell4.value = 'Сумма MinusTotal'
_cell5.value = 'Дата'
#
for descr, sum_date in dic.items():
    _cell = ws['A' + str(num)]  # Define the necessary cell on the sheet
    _cell2 = ws['B' + str(num)]
    _cell3 = ws['C' + str(num)]
    _cell4 = ws['D' + str(num)]
    _cell5 = ws['E' + str(num)]
    _cell.value = descr  # Write in the cell
    _cell2.value = ", ".join(map(lambda x: str(x).replace('.', ','), sum_date["amountsPlus"]))
    _cell3.value = ", ".join(map(lambda x: str(x).replace('.', ','), sum_date["amountsMinus"]))
    _cell4.value = ", ".join(map(lambda x: str(x).replace('.', ','), [sum_date["totalminus"]]))
    _cell5.value = str(sum_date['dates'])
    num += 1


fill_carsharing = PatternFill(start_color="FAEBD7", end_color="FAEBD7", fill_type="solid")
fill_canteen = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
fill_homefood = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
fill_Telephone_internet_metro_barbershop = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
fill_shopping_walking = PatternFill(start_color="42AAFF", end_color="42AAFF", fill_type="solid")
fill_flatrenting = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
fill_church = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

# Проходите по каждой строке и клетке на листе
for row in range(2, num):
    cell = ws["A" + str(row)]  # Определите необходимую ячейку на листе
    description = cell.value  # Получите значение ячейки

    if "CITYDRIVE" in description or "Яндекс Драйв" in description or "Белка" in description or "Делимобиль" in description:
        cell.fill = fill_carsharing
    if "Столовая" in description:
        cell.fill = fill_canteen
    if "5ка/Перекресток" in description or "Вкусвилл" in description or "Магнит" in description:
        cell.fill = fill_homefood
    if "Тинькофф мобайл" in description or "MTS" in description or "Метро" in description or "Парикмахерская" in description:
        cell.fill = fill_Telephone_internet_metro_barbershop
    if "Дикси" in description or "Wildberries" in description or "EUROSPAR" in description or "Вкусно и точка" in description or "KFC" in description or "Вендинг машина" in description:
        cell.fill = fill_shopping_walking
    if "Аренда квартиры" in description:
        cell.fill = fill_flatrenting
    if "Пожертвования" in description:
        cell.fill = fill_church
#


# Save the workbook to the specified file
wb.save(FILE_NAME)
