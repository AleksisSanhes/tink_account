import pdfplumber
import re
dic = {}
page = 0
# total = 0
plus = 0
minus = 0
with pdfplumber.open("last.pdf") as temp:
  try:
    while temp.pages[page].extract_text():
      first_page = temp.pages[page].extract_text() # выбираем стр и вытаскиваем текст
      page += 1
      sum = re.findall(r'[₽].{0,}[₽]', first_page)  # сумма
      description = re.findall(r'[₽]\s\w.{0,}', first_page)  # описание
      # print(first_page)
      for i in range(len(description)):
        description_each =description[i][2:]  # убираем ₽
        # print(description_each)
        sign = sum[i][2:3]  # знак
        # print(sign)
        number = float(sum[i][4:].replace(',', '.').replace(' ', '').replace('₽', ''))  # цифра
        if sign == '-':
          minus -= number
        else:
          plus += number
        # print(number)
        if description_each not in dic and sign == '-':
          dic[description_each] = float('-'+str(number))
        elif description_each not in dic and sign == '+':
          dic[description_each] = number
        elif description_each in dic and sign == '-':
          dic[description_each] -= number
        elif description_each in dic and sign == '+':
          dic[description_each] += number
  except IndexError:
    pass

# print(sum)
# print(plus, minus)
# for i in sum:
#
#   print(i)
# print(dic)
# print(plus)
# print(minus)

# for i in dic:
#   print(i, dic[i])
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.fills import PatternFill

# Загружаем документ
book = load_workbook('foo.xlsx')

# Определяем рабочий лист
ws = book.worksheets[0]



# Задаем стиль для ячейки
# _cell.font = Font(size=10, underline='single', color='FFFFFF', bold=True, italic=True)

num = 1
for descr, sum in dic.items():
  _cell = ws['A'+str(num)] # Определяем необходимую ячейку на листе
  _cell2 = ws['B' + str(num)]
  _cell.value = descr
  _cell2.value = sum
  num += 1
  # print(descr, '->', sum)

# Задаем цвет фона
# _cell.fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")

# Указываем ширину для колонки
# ws.column_dimensions["C"].width = 60.0

# Сохраняем документ
book.save('foo.xlsx')




