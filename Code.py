import pdfplumber
from abbrevia import abbreviation
from datetime import date  # для сравнения дат
from openpyxl.writer.excel import save_workbook
from openpyxl import *

dic = {}
page = 0
from_date = "01.04.2023".split('.')  # сюда вводи дату, с какой счет вести
from_date = date(int(from_date[2]), int(from_date[1]), int(from_date[0]))

upto_date = "01.05.2023".split('.')  # сюда вводи дату, с какой счет вести
upto_date = date(int(upto_date[2]), int(upto_date[1]), int(upto_date[0]))


def in_dic(descrip):
    if descrip in dic:
        if sign == '-':
            dic[descrip] -= float(amount)
        else:
            dic[descrip] += float(amount)
    else:
        dic[descrip] = float(sign + amount)


with pdfplumber.open("periods/1.05.2023.pdf") as temp:
    try:
        while temp.pages[page].extract_text():
            first_page = temp.pages[page].extract_text()  # текст страницы <class 'str'>
            first_page = first_page.split('\n')  # разделение по строкам <class 'list'>
            for i in first_page:
                separated = i.split()  # разделение по пробелам
                if separated[0][0].isdigit() and len(separated) > 9:  # нужные строки с датами
                    date_taken = separated[0]
                    current_date = date(int(date_taken[6:10]), int(date_taken[3:5]), int(date_taken[0:2]))
                    if from_date <= current_date and upto_date >= current_date:
                        sign = separated[3]
                        amount = separated[4] if separated[5][-1] != '₽' else separated[4] + separated[5]
                        amount = amount.replace('₽', '').replace(',', '.')
                        description = separated[8:] if not separated[8:][0][0].isdigit() else separated[9:]
                        description = description if description[0] != '₽' else description[1:]
                        description = ' '.join(description)
                        description = abbreviation(description)
                        in_dic(description)
            page += 1
    except IndexError:
        pass
# print(dic)

# ---------------------Запись в файл--------------------------#


FILE_NAME = 'excel.xlsx'

try:
    wb = load_workbook(FILE_NAME)
except:
    wb = Workbook()

ws = wb.create_sheet(str(from_date) + ' - ' + str(upto_date))

num = 2
_cell = ws['A1']
_cell2 = ws['B1']

_cell.value = 'Описание'
_cell2.value = 'Сумма'

for descr, SUM in dic.items():
    _cell = ws['A' + str(num)]  # Определяем необходимую ячейку на листе
    _cell2 = ws['B' + str(num)]

    _cell.value = descr  # Пишем в ячейке
    _cell2.value = SUM
    num += 1
    # print(descr, '->', SUM)

save_workbook(wb, FILE_NAME)
