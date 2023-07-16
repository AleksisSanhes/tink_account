from abbrevia import abbreviation
import PyPDF2
import re
from datetime import date  # для сравнения дат
from openpyxl.styles import PatternFill


# Define a function to extract text from a PDF file
def extract_text_from_pdf(file):
    pdf_reader = PyPDF2.PdfReader(file)
    text = ''
    for page_num in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[page_num]
        text += page.extract_text()
    return text

# Read the PDF file and extract its text
with open('periods/1.07.2023.pdf', 'rb') as f:
    pdf_text = extract_text_from_pdf(f)

# Define a regular expression pattern to match the amounts of money and descriptions
pattern_amount_description = r'([-+]?\d{1,3}(?:[ ,]\d{3})*(?:\.\d{2})?) ₽ [^₽]*₽ (.*[^..]*)'
pattern_date = r'(\d{2}\.\d{2}\.\d{4})\s\d{2}:\d{2}:\d{2}\s\d{2}'
# Find all matches of the pattern in the text
Match_amount_description = re.findall(pattern_amount_description, pdf_text)
matches_date = re.findall(pattern_date, pdf_text)

# -------мой код---------------------
dic = {}
total = 0

def in_dic(amounts, descrip, date):
    if descrip in dic:
        if str(amounts)[0] == '-':
            dic[descrip]['amounts'] = dic[descrip]['amounts'] + amounts
        else:
            dic[descrip]['amounts'] += amounts
        dic[descrip]['dates'].append(date)
    else:
        dic[descrip] = {'amounts': amounts, 'dates': [date]}

from_date = "01.06.2023".split('.')  # сюда вводи дату, с какой счет вести
from_date = date(int(from_date[2]), int(from_date[1]), int(from_date[0]))

upto_date = "01.07.2023".split('.')  # сюда вводи дату, с какой счет вести
upto_date = date(int(upto_date[2]), int(upto_date[1]), int(upto_date[0]))

for match in range(len(matches_date)):
    Match_amount_description_string = Match_amount_description[match]
    amount_string = float(Match_amount_description_string[0].replace(',', '').replace(' ', '')) # Extract the amounts of money from the matches and convert them to float numbers
    description_string = abbreviation(Match_amount_description_string[1][:-3])  # Extract the descriptions from the matches
    date_string = matches_date[match]
    date_int = date_string.split('.')
    date_int = date(int(date_int[2]), int(date_int[1]), int(date_int[0]))
    if from_date <= date_int < upto_date:
        in_dic(amount_string, description_string, date_string)
        total += amount_string # считает правильно получается

print(dic)
print(total)
#
# ---------------------Запись в файл--------------------------#

from openpyxl.writer.excel import save_workbook
from openpyxl import *

page = 0

FILE_NAME = 'excel.xlsx'

try:
    wb = load_workbook(FILE_NAME)
except:
    wb = Workbook()

ws = wb.create_sheet(str(from_date) + ' - ' + str(upto_date))

num = 2
_cell = ws['A1']
_cell2 = ws['B1']
_cell3 = ws['C1']

_cell.value = 'Описание'
_cell2.value = 'Сумма'
_cell3.value = 'Дата'

for descr, sum_date in dic.items():
    _cell = ws['A' + str(num)]  # Определяем необходимую ячейку на листе
    _cell2 = ws['B' + str(num)]
    _cell3 = ws['C' + str(num)]
    _cell.value = descr  # Пишем в ячейке
    _cell2.value = sum_date["amounts"]
    _cell3.value = str(sum_date['dates'])
    num += 1
    # print(descr, '->', SUM)




fill_carsharing = PatternFill(start_color="FAEBD7", end_color="FAEBD7", fill_type="solid")
fill_canteen = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
fill_homefood = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
fill_Telephone_internet_metro_barbershop = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
fill_shopping_walking = PatternFill(start_color="42AAFF", end_color="42AAFF", fill_type="solid")
fill_flatrenting = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
fill_church = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

# Проходите по каждой строке и клетке на листе
for row in range(2, num):
    cell = ws["A" + str(row)]  # Определите необходимую ячейку на листе
    description = cell.value  # Получите значение ячейки

    if "CITYDRIVE" in description or "Яндекс Драйв" in description or "Белка" in description or "Делимобиль" in description:
        cell.fill = fill_carsharing
    if "Столовая" in description:
        cell.fill = fill_canteen
    if "Пятерочка" in description or "5ка/Перекресток" in description or "Вкусвилл" in description or "Магнит" in description:
        cell.fill = fill_homefood
    if "Тинькофф мобайл" in description or "MTS" in description or "Метро" in description or "Парикмахерская" in description:
        cell.fill = fill_Telephone_internet_metro_barbershop
    if "Дикси" in description or "Wildberries" in description or "EUROSPAR" in description or "Вкусно и точка" in description or "KFC" in description or "Вендинг машина" in description:
        cell.fill = fill_shopping_walking
    if "Аренда квартиры" in description:
        cell.fill = fill_flatrenting
    if "Пожертвования" in description:
        cell.fill = fill_church

save_workbook(wb, FILE_NAME)


